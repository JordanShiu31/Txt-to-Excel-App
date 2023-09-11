import glob
import openpyxl as op
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from functools import partial
from itertools import product
import shutil

#-----------------------------------------------------#
# ---------- Create the application window ---------- #
#-----------------------------------------------------#

window = tk.Tk()
window.title("Transferring Txt data into Excel")
window.geometry('1080x720')
window.minsize(1080,720)

# styles for placeholder text
style = ttk.Style()
style.configure("Gray.TEntry", foreground="#666")
style.configure("Black.TEntry", foreground="black")
selected_worksheet_names = []
scenario_state = []
scenario_years = []
scenario_period = []

#----------------------------------------------------#
# ---------- Functions to operate the GUI ---------- #
#----------------------------------------------------#

## Functions for TEXTBOX Widget
def list_from_comma_separated_str(input_word):
    return [word.strip() for word in input_word.split(',') if word]

def update_txt_files(curr_listbox, functionality):
    root_directory_path = txt_root_directory_entry.get().strip().strip('"\'').strip()
    common_file_name = txt_result_file_names_entry.get().strip()
    existing_words = list_from_comma_separated_str(txt_existing_name_entry.get().strip())
    replacing_words = list_from_comma_separated_str(txt_replacing_name_entry.get().strip())
    destination_path = txt_final_folder_path_entry.get().strip().strip('"\'').strip()

    if not os.path.exists(root_directory_path):
        messagebox.showerror("Error", "Incorrect Folder Directory!\nPlease check the entered path name")
        return

    # convert existing/replacing words into key/value dictionary
    if len(existing_words) == len(replacing_words) and functionality == "UPDATE TXT":
        replacing_dict = {existing: replacing for existing, replacing in zip(existing_words, replacing_words)}
    elif functionality != "SHOW TXT":
        messagebox.showerror("Error", f"The number of desired words to replace ({len(existing_words)}) is not equal to the number of replacing words ({len(replacing_words)})!")
        return
    
    txt_files_dict = get_common_ending_filenames_in_root_recursive(root_directory_path, common_file_name)

    if txt_files_dict is None or len(txt_files_dict['file_paths']) == 0:
        messagebox.showerror("Error", f"No files ending with '{common_file_name}' were found in {root_directory_path}.\nPlease check the common ending name or folder path!")
    elif functionality == "SHOW TXT":
        txt_show_updated_button.config(state=tk.NORMAL)
        update_listbox(txt_files_dict['file_names'], curr_listbox)
    elif functionality == "UPDATE TXT":
        # one liner: for each word in replacing_dict, if the word is not found in any filename, add it to words_not_found
        if words_not_found := [word for word in replacing_dict if not any([word for filename in txt_files_dict['file_names'] if word.upper() in filename.upper()])]:
            messagebox.showerror("Error", f"The following existing word(s) cannot be found in any filenames:\n{', '.join(w for w in words_not_found)}")
            return

        txt_delete_new_files_button.config(state=tk.NORMAL)

        files_updated = 0
        for path, filename in zip(txt_files_dict['file_paths'], txt_files_dict['file_names']):
            for existing, replacing in replacing_dict.items():
                if existing.upper() in filename.upper():
                    filename = filename.replace(existing, replacing)
            final_destination_path = os.path.join(destination_path, filename).upper()
            shutil.copy(path, final_destination_path)
            files_updated += 1
            
        new_file_list = get_common_ending_filenames_in_root_recursive(destination_path, common_file_name)['dest_files']
        update_listbox(new_file_list, curr_listbox)
        messagebox.showinfo("Success!",f"Updated {files_updated} file{'s' if files_updated != 1 else ''}!")

## Removes white spaces and separates each line to a list
def textbox_to_list(textbox, button):
    btn_text: str = button.cget('text')
    formatted_data: str = textbox.get('1.0', 'end').replace(' ', '').strip()

    if formatted_data == '':
        button.config(text=f"Please enter a {'year' if 'year' in btn_text else 'state'}(s)")
        textbox.delete('1.0', 'end')
        return

    data_list = formatted_data.split('\n')
    if 'year' in btn_text:
        global scenario_years
        scenario_years = data_list
        button.config(text=f"{len(data_list)} years selected")
    else:
        global scenario_state
        scenario_state = data_list
        button.config(text=f"{len(data_list)} states selected")

#-------------------------------------------------------------------------------------------#

#-------------------------------------------------------------------------------------------#
def delete_selected_listbox_items(curr_listbox):
    destination_path = destination_path_stringVar.get()
    selected_items = get_selected_listbox_items(curr_listbox)
    for file in selected_items:
        try:
            os.remove(os.path.join(destination_path, file))
        except FileNotFoundError:
            continue

    selected_indices = curr_listbox.curselection()
    for index in selected_indices[::-1]:
        curr_listbox.delete(index)
    if not [file for file in os.listdir(destination_path) if os.path.isfile(os.path.join(destination_path, file))]:
        txt_delete_new_files_button.config(state=tk.DISABLED)

## Returns a list of selected tab names
def get_selected_listbox_items(listbox):
    return [listbox.get(index) for index in listbox.curselection()]

## Updates the number of periods selected and the button text
def assign_scenario_period_to_list(curr_listbox, button):
    global scenario_period
    scenario_period = get_selected_listbox_items(curr_listbox)
    button.config(text=f"{len(scenario_period)} periods selected")
#-------------------------------------------------------------------------------------------#

#-------------------------------------------------------------------------------------------#
## Functions for EXCEL LISTBOX Widget
## Updates the number of worksheets selected and the button text
def set_num_selected_worksheets(worksheet_listbox, select_worksheets_button):
    selected_worksheet_names = get_selected_listbox_items(worksheet_listbox)
    select_worksheets_button.config(text=f"{len(selected_worksheet_names)} worksheets selected")
    update_excel_button.config(state=tk.NORMAL) # unlocks the update button

## Adds existing worksheet tabs into the listbox    
def update_listbox(items_list, listbox):
    listbox.delete(0, tk.END)
    listbox.insert(tk.END, *items_list)

## deletes all data in the listbox
# attempts to open user input excel file and will display errors if excel file is not valid
# adds all existing excel tabs (worksheets) into the list
def get_worksheet_names(excel_stringvar, listbox):
    # converts to a string
    file_path = excel_stringvar.get().strip().strip('"\'').strip()
    try:
        workbook = op.load_workbook(file_path)
        update_listbox(workbook.sheetnames, listbox)
        workbook.close()
        excel_path_status_msg.set(value="Valid Spreadsheet")
        select_worksheet_button.config(state=tk.NORMAL) # NEED TO ADD IF PARAMS ARE SET AS WELL
    except op.utils.exceptions.InvalidFileException:
        excel_path_status_msg.set(value="Error: Invalid path, please enter a correct excel path")
    except FileNotFoundError:
        excel_path_status_msg.set(value="Error: Excel file not found")
#-------------------------------------------------------------------------------------------#

## Returns a list of keywords that matches the user input and user desired worksheet names
def find_worksheet_key_words(worksheet_name):
    desired_words = []
    # product returns generator for all possible permutations of states, years and periods
    for perm in product(scenario_state, scenario_period, scenario_years):
        all_in_worksheet_name = True
        for item in perm:
            if item.upper() not in worksheet_name.upper():
                all_in_worksheet_name = False
        if all_in_worksheet_name:
            desired_words.extend(perm)
    return desired_words

# Returns a list containing the FULL txt file paths and the Txt file names with common ending
# check_same_dir = False for when root_directory and destination_path are the same folder path
def get_common_ending_filenames_in_root_recursive(root_directory, common_ending_file_name, check_same_dir=True):
    destination_path = destination_path_stringVar.get()
    file_pattern = os.path.join(root_directory.replace('\\', '/'), f"**/*{common_ending_file_name}")
    files_in_root = [file.replace('\\', '/') for file in glob.glob(file_pattern, recursive=True)]
    if check_same_dir:
        files_in_root = [file for file in files_in_root if not os.path.dirname(file).replace('\\', '/') == destination_path.replace('\\', '/')]

    list_raw_dest_files = []   
    try:
        for file in os.listdir(destination_path):
            if os.path.isfile(os.path.join(destination_path, file)):
                list_raw_dest_files.append(file)
    except OSError:
        messagebox.showerror("Error", f"The path: {destination_path} does not exist!\nPlease enter an existing folder")
        return
    
    return {
        'file_paths': files_in_root,
        'file_names': [os.path.basename(file) for file in files_in_root],
        'dest_files': list_raw_dest_files,
    }

# Returns a list of the full txt file paths which contain the desired keywords
def find_txt_file(keywords_list, raw_data_files_dict): 
    for index, path_name in enumerate(raw_data_files_dict['file_names']):
        counter = 0
        for word in keywords_list:
            if word in path_name:
                counter += 1
        if counter == len(keywords_list):
            return raw_data_files_dict['file_paths'][index]
    return None        

## Opens the selected full txt file path and reads all lines, clears all existing excel contents and pastes new contents
def open_txt_file(desired_path_file, delimiter, worksheet_active):
    with open(desired_path_file, 'r') as file:
        lines = file.readlines()
    clear_worksheet_contents(worksheet_active)
    paste_contents(lines, delimiter, worksheet_active)
        
## clears all exsisting data inside the active worksheet
def clear_worksheet_contents(worksheet_active):
    for row in worksheet_active.iter_rows():
        for cell in row:
            cell.value = None

## pastes each line of txt data into the active worksheet and attempts to paste them as int value type
def paste_contents(lines, delimiter, worksheet):
    starting_row = 1
    starting_column = 1
    for row_idx, line in enumerate(lines, start=starting_row):
        values = line.strip().split(delimiter)  # Replace ',' with your desired delimiter
        for col_idx, value in enumerate(values, start=starting_column):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
            if value.isdigit():
                worksheet.cell(row=row_idx, column=col_idx, value=int(value))
            else:
                try:
                    worksheet.cell(row=row_idx, column=col_idx, value=float(value))
                except ValueError:
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

# the main function which does the magic when the update excel button is pressed
def update_excel(curr_listbox):
    excel_path = excel_stringVar.get().strip().strip('"\'').strip()
    root_directory_path = root_directory_entry.get().strip().strip('"\'').strip()
    common_ending_file_name = txt_result_file_names_string.get()
    delimiter = delimiter_entry.get()
    
    if not os.path.exists(root_directory_path):
        messagebox.showerror("Error", "Incorrect Folder Directory!\nPlease check the entered path name")
        return
    
    # if the root folder is found but no result files could be found it could be an error
    raw_data_files_dict = get_common_ending_filenames_in_root_recursive(root_directory_path, common_ending_file_name, check_same_dir=False)
    if raw_data_files_dict is None or len(raw_data_files_dict['file_paths']) == 0:
        show_error_message_box(raw_data_files_dict['file_paths'], root_directory_path, common_ending_file_name)
        return
    
    unselected_worksheets = []
    pasted_contents_worksheets = []
    workbook = op.load_workbook(excel_path)
    for worksheet_name in get_selected_listbox_items(curr_listbox):
        worksheet_active = workbook[worksheet_name]
        if len(keywords_list := find_worksheet_key_words(worksheet_name)) < 3:
            unselected_worksheets.append(worksheet_name)
            continue     
        if desired_path_file := find_txt_file(keywords_list, raw_data_files_dict):
            open_txt_file(desired_path_file, delimiter, worksheet_active)
            pasted_contents_worksheets.append(worksheet_name)
        else:
            messagebox.showerror("Error", f"raw data file could not be found! for:\n {worksheet_name}")
            unselected_worksheets.append(worksheet_name)
            
    try:
        workbook.save(excel_path)
    except PermissionError:
        messagebox.showerror("Error", "The excel file is still open!\nPlease close it first, then press the update button")
        workbook.close()
        return      

    workbook.close()
    messagebox.showinfo(
        "Complete",
        (
            f"Data has been pasted in Excel for these scenarios:\n"
            f"{', '.join(w for w in pasted_contents_worksheets)}\n"
            "However, the below scenarios were not pasted due to user scenario parameters:\n"
            f"{', '.join(w for w in unselected_worksheets)}\n"
        )    
    )

## Pop up errorbox for no valid files found in folder
def show_error_message_box(raw_data_files_list, root_directory_path, common_ending_file_name):
    messagebox.showerror(
        "Error",
        (
            f"There were {len(raw_data_files_list)} found files found in folder "
            f"with keywords :\n{common_ending_file_name}\n\n"
            f"Please double check the 'Raw Results File Name' or there may be "
            f"no results at all inside:\n{root_directory_path}"
        )
    )

def browse_excel():
    if file_path := filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls .xlsm")]):
       excel_stringVar.set(file_path)
       excel_path_entry.config(style="Black.TEntry")

def browse_root():
    if folder_path := filedialog.askdirectory():
        txt_root_directory_string.set(folder_path)
        txt_root_directory_entry.config(style="Black.TEntry")

def browse_destination():
    if folder_path := filedialog.askdirectory():
        destination_path_stringVar.set(folder_path)
        root_directory_entry.config(style="Black.TEntry")
        txt_final_folder_path_entry.config(style="Black.TEntry")

def handle_FocusIn(entry: ttk.Entry, stringVar: tk.StringVar, placeholder: str, event):
    if stringVar.get() == placeholder:
        stringVar.set('')
        entry.config(style="Black.TEntry")
        if entry == txt_result_file_names_entry:
            result_file_names_entry.config(style="Black.TEntry")
        elif entry == result_file_names_entry:
            txt_result_file_names_entry.config(style="Black.TEntry")
        elif entry == txt_final_folder_path_entry:
            root_directory_entry.config(style="Black.TEntry")
        elif entry == root_directory_entry:
            txt_final_folder_path_entry.config(style="Black.TEntry")

def handle_FocusOut(entry: ttk.Entry, stringVar: tk.StringVar, placeholder: str, event):
    if stringVar.get() == '':
        stringVar.set(placeholder)
        entry.config(style="Gray.TEntry")
        if entry == txt_result_file_names_entry:
            result_file_names_entry.config(style="Gray.TEntry")
        elif entry == result_file_names_entry:
            txt_result_file_names_entry.config(style="Gray.TEntry")
        elif entry == txt_final_folder_path_entry:
            root_directory_entry.config(style="Gray.TEntry")
        elif entry == root_directory_entry:
            txt_final_folder_path_entry.config(style="Gray.TEntry")

def on_tab_selected(_):
    # when switching tabs, the first entry/textbox won't be highlighted
    window.focus()

#-----------------------------------------------#
# ---------- Creating all the widgets --------- #
#-----------------------------------------------#

notebook_tabs = ttk.Notebook(window)
notebook_tabs.bind("<<NotebookTabChanged>>", on_tab_selected)
notebook_tabs.pack(expand=1, fill="both")

#-----------------------------------------------------#
# ---------- Creating txt renaming widgets ---------- #
#-----------------------------------------------------#

# 1. Creates a frame inside window to allow some padding and to help with widget placement
txt_renaming_frame = tk.Frame(notebook_tabs)
txt_renaming_frame.pack(expand=True, fill="both", padx=10, pady=10) 
notebook_tabs.add(txt_renaming_frame, text="Txt File Renaming")

# 2. Creates subframes inside frame above to allow grouping of other widgets
txt_input_frame = tk.Frame(txt_renaming_frame)

# Label frame and widgets
txt_input_labels_frame = tk.Frame(txt_input_frame)
txt_root_directory_label = ttk.Label(txt_input_labels_frame, text="Full Folder Path")
txt_result_file_names_label = ttk.Label(txt_input_labels_frame, text="Common Txt File name")
txt_existing_name_label = ttk.Label(txt_input_labels_frame, text="Desired word to replace")
txt_replacing_name_label = ttk.Label(txt_input_labels_frame, text="New word to replace old word")
txt_final_folder_path_label = ttk.Label(txt_input_labels_frame, text="Destination Folder")

# Entries frame and widgets
txt_input_entries_frame = tk.Frame(txt_input_frame)
entry_width = 100
txt_root_directory_string = tk.StringVar(value="Please browse or enter the full folder path file containing all result files")
txt_root_directory_entry = ttk.Entry(txt_input_entries_frame, style="Gray.TEntry", textvariable=txt_root_directory_string, width=entry_width)
txt_root_dir_browse_button = ttk.Button(txt_input_entries_frame, text="Browse Folder", width=20, command=lambda:browse_root())

# txt_result_file_names_string shared between the renaming and excel tabs
txt_result_file_names_string = tk.StringVar(value="Please enter a common results file name e.g. Vehicle Travel Time Results.att")
txt_result_file_names_entry = ttk.Entry(txt_input_entries_frame, style="Gray.TEntry", textvariable=txt_result_file_names_string, width=entry_width)

txt_existing_name_string = tk.StringVar(value="Please enter word(s) to replace in txt filenames. Separate each word with a comma (,)")
txt_existing_name_entry = ttk.Entry(txt_input_entries_frame, style="Gray.TEntry", textvariable=txt_existing_name_string, width=entry_width)

txt_replacing_name_string = tk.StringVar(value="Please enter word(s) to replace existing word(s) in txt filenames. Corresponds with position of each word in previous textbox.")
txt_replacing_name_entry = ttk.Entry(txt_input_entries_frame, style="Gray.TEntry", textvariable=txt_replacing_name_string, width=entry_width)

# destination_path_stringVar shared between the renaming and excel tabs
destination_path_stringVar = tk.StringVar(value="Please browse or enter a destination folder path")
txt_final_folder_path_entry = ttk.Entry(txt_input_entries_frame, style="Gray.TEntry", textvariable=destination_path_stringVar, width=entry_width)
txt_dest_dir_browse_button = ttk.Button(txt_input_entries_frame, text="Browse Destination", width=20, command=lambda:browse_destination())

# Bind focus in/out events to entries
for entry, stringVar, placeholder in [
        (txt_root_directory_entry, txt_root_directory_string, "Please browse or enter the full folder path file containing all result files"),
        (txt_result_file_names_entry, txt_result_file_names_string, "Please enter a common results file name e.g. Vehicle Travel Time Results.att"),
        (txt_existing_name_entry, txt_existing_name_string, "Please enter word(s) to replace in txt filenames. Separate each word with a comma (,)"),
        (txt_replacing_name_entry, txt_replacing_name_string, "Please enter word(s) to replace existing word(s) in txt filenames. Corresponds with position of each word in previous textbox."),
        (txt_final_folder_path_entry, destination_path_stringVar, "Please browse or enter a destination folder path")
    ]:
    entry.bind("<FocusIn>", partial(handle_FocusIn, entry, stringVar, placeholder))
    entry.bind("<FocusOut>", partial(handle_FocusOut, entry, stringVar, placeholder))

# Display frames and widgets
txt_display_frame = tk.Frame(txt_renaming_frame)
txt_display_existing_frame = tk.Frame(txt_display_frame)
txt_existing_error_msg = tk.StringVar(value="Existing files (click button to verify)")
txt_existing_file_label = ttk.Label(txt_display_existing_frame, textvariable=txt_existing_error_msg)
txt_show_existing_button = ttk.Button(txt_display_existing_frame, text="Show existing txt file names", command=lambda:update_txt_files(txt_existing_path_name_listbox, functionality="SHOW TXT"))
txt_existing_path_name_listbox = tk.Listbox(txt_display_existing_frame)

txt_display_updated_frame = tk.Frame(txt_display_frame)
txt_updated_error_msg = tk.StringVar(value="Updated files (click below to update)")
txt_updated_file_label = ttk.Label(txt_display_updated_frame, textvariable=txt_updated_error_msg)
txt_updated_path_name_listbox = tk.Listbox(txt_display_updated_frame, selectmode=tk.EXTENDED)
txt_updated_buttons_frame = tk.Frame(txt_display_updated_frame)
txt_show_updated_button = ttk.Button(txt_updated_buttons_frame, text="Update txt file names", command=lambda:update_txt_files(txt_updated_path_name_listbox, functionality="UPDATE TXT"), state=tk.DISABLED)
txt_delete_new_files_button = ttk.Button(txt_updated_buttons_frame, text="Delete selected txt files", command=lambda:delete_selected_listbox_items(txt_updated_path_name_listbox), state=tk.DISABLED)

#----------------------------------------------------------------------#
# ---------- Placing all renaming widgets into renaming tab ---------- #
#----------------------------------------------------------------------#
xpad = 10
ypad = 20

txt_input_frame.pack(side='top', fill="both", expand=False, padx=xpad, pady=10)
txt_display_frame.pack(side='top', fill="both", expand=True, padx=xpad)
txt_display_frame.grid_rowconfigure(0, weight=1)
txt_display_frame.grid_columnconfigure(0, weight=4)
txt_display_frame.grid_columnconfigure(1, weight=3)
txt_input_labels_frame.pack(side='left', fill="both", expand=False, padx=xpad, pady=ypad)
txt_input_entries_frame.pack(side='left', fill="both", expand=True, padx=xpad, pady=ypad)
txt_display_existing_frame.grid(row=0, column=0, sticky="nsew", padx=xpad, pady=ypad)
txt_display_updated_frame.grid(row=0, column=1, sticky="nsew", padx=xpad, pady=ypad)

# Input frame widgets
txt_root_directory_label.pack(side='top', fill="both", expand=True, padx=xpad)
txt_result_file_names_label.pack(side='top', fill="both", expand=True, padx=xpad)
txt_existing_name_label.pack(side='top', fill="both", expand=True, padx=xpad)
txt_replacing_name_label.pack(side='top', fill="both", expand=True, padx=xpad)
txt_final_folder_path_label.pack(side='top', fill="both", expand=True, padx=xpad)

txt_root_directory_entry.pack(side='top', fill="both", expand=True, padx=xpad, pady=(ypad, 0))
txt_root_dir_browse_button.pack(anchor='w', side='top', expand=False, padx=xpad, pady=(1, 0))

txt_result_file_names_entry.pack(side='top', fill="both", expand=True, padx=xpad, pady=(ypad-1, ypad))
txt_existing_name_entry.pack(side='top', fill="both", expand=True, padx=xpad, pady=ypad)
txt_replacing_name_entry.pack(side='top', fill="both", expand=True, padx=xpad, pady=ypad)
txt_final_folder_path_entry.pack(side='top', fill="both", expand=True, padx=xpad, pady=(ypad, 0))
txt_dest_dir_browse_button.pack(anchor='w', side='top', expand=False, padx=xpad, pady=(1, 0))

# Display frame widgets
txt_existing_file_label.pack(side='top', fill="both", expand=False, padx=xpad)
txt_existing_file_label.config(anchor="center")
txt_updated_file_label.pack(side='top', fill="both", expand=False, padx=xpad)
txt_updated_file_label.config(anchor="center")
txt_existing_path_name_listbox.pack(side='top', fill="both", expand=True, padx=xpad, pady=(5, 1))
txt_updated_path_name_listbox.pack(side='top', fill="both", expand=True, padx=xpad, pady=(5, 0))
txt_show_existing_button.pack(side='top', fill="both", expand=False, padx=xpad, pady=1)

txt_updated_buttons_frame.pack(fill="both", expand=False, pady=0)
txt_show_updated_button.pack(side='left', fill="x", expand=True, padx=(xpad, 3), pady=1)
txt_delete_new_files_button.pack(side='right', fill="x", expand=True, padx=(3, xpad))

#----------------------------------------------#
# ---------- Creating Excel Widgets ---------- #
#----------------------------------------------#

# Create subframes inside frame above to allow grouping of other widgets
excel_frame = tk.Frame(notebook_tabs)
excel_frame.pack(expand=True, fill="both", padx=10, pady=10)
notebook_tabs.add(excel_frame, text="Txt to Excel")

input_frame = tk.LabelFrame(excel_frame, text="Inputs")

step_1_and_2_frame = tk.LabelFrame(input_frame, text="[STEP 1.] File Inputs", width=300, height=400)
file_input_frame = tk.Frame(step_1_and_2_frame)

#--------------------------------------------#
# ---------- Excel Step 1 widgets ---------- #
#--------------------------------------------#

# Input labels frame and widgets
file_input_label_frame = tk.Frame(file_input_frame)
excel_path_label = ttk.Label(file_input_label_frame, text="Full Excel Path")
root_directory_label = ttk.Label(file_input_label_frame, text="Full Folder Path")
result_file_names_label = ttk.Label(file_input_label_frame, text="Raw Results File Name")
delimiter_label = ttk.Label(file_input_label_frame, text="Delimiter")

# Input entries frame and widgets
file_input_entries_frame = tk.Frame(file_input_frame)
entry_width = 50

excel_stringVar = tk.StringVar(value="Enter full excel file path")
excel_path_entry = ttk.Entry(file_input_entries_frame, style="Gray.TEntry", textvariable=excel_stringVar, width=entry_width)
browse_excel_button = ttk.Button(file_input_entries_frame, text="Browse Excel", width=20, padding=0, command=lambda:browse_excel())

root_directory_entry = ttk.Entry(file_input_entries_frame, style="Gray.TEntry", textvariable=destination_path_stringVar, width=entry_width)
excel_browse_destination_button = ttk.Button(file_input_entries_frame, text="Browse Destination", width=20, padding=0, command=lambda:browse_destination())

result_file_names_entry = ttk.Entry(file_input_entries_frame, style="Gray.TEntry", textvariable=txt_result_file_names_string, width=entry_width)

delimiter_stringVar = tk.StringVar(value="Enter a delimiter to separate the results")
delimiter_entry = ttk.Entry(file_input_entries_frame, style="Gray.TEntry", textvariable=delimiter_stringVar, width=entry_width)

# Bind focus in/out events to entries
for entry, stringVar, placeholder in [
        (excel_path_entry, excel_stringVar, "Enter full excel file path"),
        (root_directory_entry, destination_path_stringVar, "Please browse or enter a destination folder path"),
        (result_file_names_entry, txt_result_file_names_string, "Please enter a common results file name e.g. Vehicle Travel Time Results.att"),
        (delimiter_entry, delimiter_stringVar, "Enter a delimiter to separate the results")
    ]:
    entry.bind("<FocusIn>", partial(handle_FocusIn, entry, stringVar, placeholder))
    entry.bind("<FocusOut>", partial(handle_FocusOut, entry, stringVar, placeholder))

#--------------------------------------------#
# ---------- Excel Step 2 widgets ---------- #
#--------------------------------------------#
scenario_frame = tk.LabelFrame(step_1_and_2_frame, text="[STEP 2.] Scenario Parameters (Please select keywords from txt file name)")

scenario_years_frame = tk.Frame(scenario_frame)
scenario_years_label = tk.Label(scenario_years_frame, text="Select Scenario Years")
scenario_years_input_text = tk.Text(scenario_years_frame, height=5, width=15)
scenario_years_button = ttk.Button(scenario_years_frame, text="Select years", command=lambda:textbox_to_list(scenario_years_input_text, scenario_years_button))

scenario_state_frame = tk.Frame(scenario_frame)
scenario_state_label = tk.Label(scenario_state_frame, text="Select Scenario State")
scenario_state_input_text = tk.Text(scenario_state_frame, height=5, width=15)
scenario_state_button = ttk.Button(scenario_state_frame, text="Select state", command=lambda:textbox_to_list(scenario_state_input_text, scenario_state_button))

scenario_period_frame = tk.Frame(scenario_frame)
scenario_period_label = tk.Label(scenario_period_frame, text="Select Scenario Period")
scenario_period_input_listbox = tk.Listbox(scenario_period_frame, selectmode="multiple", height=5)
scenario_period_button = ttk.Button(scenario_period_frame, text="Select periods", command=lambda:assign_scenario_period_to_list(scenario_period_input_listbox, scenario_period_button))

#--------------------------------------------#
# ---------- Excel Step 3 widgets ---------- #
#--------------------------------------------#
excel_tab_list_frame = tk.LabelFrame(input_frame, text="[STEP 3.] List of Excel Tabs (Worksheets)")
excel_path_status_msg = tk.StringVar(value="Click Show Worksheets to verify valid path") # stringVar status message that updates when 'Show Worksheets' button is clicked
excel_path_status_label = ttk.Label(excel_tab_list_frame, textvariable=excel_path_status_msg)
show_worksheets_button = ttk.Button(excel_tab_list_frame, text="Show Worksheets", command=lambda:get_worksheet_names(excel_stringVar, excel_worksheet_listbox))
excel_worksheet_listbox = tk.Listbox(excel_tab_list_frame, selectmode="extended")
worksheet_reminder_label = ttk.Label(excel_tab_list_frame, text="Please select all parameters and worksheets\nbefore pressing the 'Select Worksheets' button below")
select_worksheet_button = ttk.Button(excel_tab_list_frame, text="Select Workheets", command=lambda:set_num_selected_worksheets(excel_worksheet_listbox, select_worksheet_button), state=tk.DISABLED)

#--------------------------------------------#
# ---------- Excel Step 4 widgets ---------- #
#--------------------------------------------#
output_frame = tk.LabelFrame(excel_frame, text="[STEP 4.] Output")
update_excel_button = ttk.Button(output_frame, text="Update Excel", command=lambda:update_excel(excel_worksheet_listbox), state=tk.DISABLED)

#----------------------------------------------------------------#
# ---------- Placing all excel widgets into excel tab ---------- #
#----------------------------------------------------------------#

# 1. Place the frames into the window
xpad = 10
ypad = 20
# main input and output frames
input_frame.pack(side='top', fill="both", expand=True, padx=xpad, pady=10)
output_frame.pack(side='top', fill="both", expand=False, padx=xpad, pady=10)
# two sub frames inside the input frame
step_1_and_2_frame.pack(side='left', fill="both", expand=True, padx=xpad, pady=10)
step_1_and_2_frame.pack_propagate(False) # stops the frames from scaling
excel_tab_list_frame.pack(side='left', fill="both", expand=True, padx=xpad, pady=10)
excel_tab_list_frame.pack_propagate(False)
# Frames inside the "File Inputs" frame
file_input_frame.pack(side='top', fill="both", expand=False)
scenario_frame.pack(side='top', fill="both", expand=True)
file_input_label_frame.pack(side='left', fill="both", expand=False, padx=xpad, pady=ypad)
file_input_entries_frame.pack(side='left', fill="both", expand=True, padx=xpad, pady=ypad)

# Frames inside the 'Scenario Parameters' frame
scenario_years_frame.pack(side='left', fill="both", expand=True, padx=xpad)
scenario_state_frame.pack(side='left', fill="both", expand=True, padx=xpad)
scenario_period_frame.pack(side='left', fill="both", expand=True, padx=xpad)

# 2. Place the file input labels into the label frame
excel_path_label.pack(side='top', fill="both", pady=ypad)
root_directory_label.pack(side='top', fill="both", pady=ypad)
result_file_names_label.pack(side='top', fill="both", pady=ypad)
delimiter_label.pack(side='top', fill="both", pady=ypad)

# 3. Place the file input entries and browse buttons into the entry frame
excel_path_entry.pack(side='top', fill="both", pady=(ypad, 0))
browse_excel_button.pack(anchor='w', side='top', expand=False, pady=(1, 0))
root_directory_entry.pack(side='top', fill="both", pady=(9, 0))
excel_browse_destination_button.pack(anchor='w', side='top', expand=False, pady=(1, 0))
result_file_names_entry.pack(side='top', fill="both", pady=(ypad-1, ypad))
delimiter_entry.pack(side='top', fill="both", pady=(ypad-3, ypad))

# 4. Place the widgets into the scenario param frame
scenario_years_label.pack(side='top', pady=(0, ypad))
scenario_state_label.pack(side='top', pady=(0, ypad))
scenario_period_label.pack(side='top', pady=(0, ypad), anchor=tk.N)

scenario_years_input_text.pack(side='top', fill="both", expand=True)
scenario_state_input_text.pack(side='top', fill="both", expand=True)

scenario_period_input_listbox.pack(side='top', fill="both", expand=True)
scenario_period_input_listbox.insert(1, "AM")
scenario_period_input_listbox.insert(1, "PM")

scenario_years_button.pack(side='top', fill="both", pady=(ypad + 5, 5))
scenario_state_button.pack(side='top', fill="both", pady=(ypad + 5, 5))
scenario_period_button.pack(side='top', fill="both", pady=(ypad + 5, 5))

# 5. Place the excel spreadsheet related widgets into the excel tab list frame
excel_path_status_label.pack(side='top', fill="both", pady=9)
excel_path_status_label.config(anchor="center")
show_worksheets_button.pack(side='top', fill="both", padx=xpad)
excel_worksheet_listbox.pack(side='top', fill="both", expand=True, padx=xpad, pady=5)
worksheet_reminder_label.pack(side='top', fill="both", pady=0)
worksheet_reminder_label.config(anchor="center")
select_worksheet_button.pack(side='top', fill="both", padx=xpad, pady=10)

# 6. Place the 'Update Excel' button into the output frame
update_excel_button.pack(side='top', fill="both", expand=True, padx=xpad, pady=ypad)

#---------------------------------------#
# ---------- Runs the window ---------- #
#---------------------------------------#

window.mainloop()